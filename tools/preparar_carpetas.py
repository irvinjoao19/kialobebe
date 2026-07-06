#!/usr/bin/env python3
"""
Prepara una carpeta por producto para arrastrar las fotos (Kialo bebé).

Lee el Excel y crea, dentro de la carpeta de fotos, una subcarpeta por producto
ya nombrada (ej. "BOINA - Boina con lazo/"). La dueña solo arrastra las fotos de
cada producto a su carpeta — con cualquier nombre, sin renombrar nada.

Uso:
    python3 preparar_carpetas.py                 -> crea en tools/fotos_nuevas/
    python3 preparar_carpetas.py /ruta/a/Fotos   -> crea en esa carpeta (ej. Google Drive)
    python3 preparar_carpetas.py /ruta/Fotos otro.xlsx

Luego, cuando estén las fotos, corre:  python3 procesar_fotos.py [/ruta/a/Fotos]
"""
import os, sys, re, unicodedata
from openpyxl import load_workbook
from convertir import slug_codigo  # reutiliza la MISMA lógica de código que el conversor


def nombre_seguro(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Za-z0-9 ]", "", s).strip() or "producto"


def carpeta_base(args):
    for a in args:
        if not a.lower().endswith(".xlsx"):
            return a
    return "fotos_nuevas" if os.path.basename(os.getcwd()) == "tools" else os.path.join("tools", "fotos_nuevas")


def main():
    args = sys.argv[1:]
    xlsx = next((a for a in args if a.lower().endswith(".xlsx")), "Kialo_Catalogo.xlsx")
    base = carpeta_base(args)
    os.makedirs(base, exist_ok=True)

    ws = load_workbook(xlsx, data_only=True)["Productos"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    iNom = next((i for i, h in enumerate(headers) if h.startswith("nombre")), 0)

    usados, indice, creadas = set(), [], 0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=iNom + 1).value
        nombre = str(val).strip() if val else ""
        if not nombre:
            continue
        cod = b = slug_codigo(nombre); k = 2
        while cod in usados:
            cod = f"{b}{k}"; k += 1
        usados.add(cod)
        carpeta = os.path.join(base, f"{cod} - {nombre_seguro(nombre)}")
        if not os.path.isdir(carpeta):
            os.makedirs(carpeta); creadas += 1
        indice.append(f"{cod}\t{nombre}")

    with open(os.path.join(base, "_indice.txt"), "w", encoding="utf-8") as f:
        f.write("Carpetas de fotos de Kialo bebé\n")
        f.write("Arrastra las fotos de cada producto a su carpeta (cualquier nombre).\n\n")
        f.write("CÓDIGO\tPRODUCTO\n")
        f.write("\n".join(indice))

    print(f"✅ {creadas} carpeta(s) nueva(s). Total de productos: {len(indice)}")
    print(f"   Carpeta: {os.path.abspath(base)}")
    print("   1) Arrastra las fotos de cada producto a su carpeta.")
    print("   2) Luego corre:  python3 procesar_fotos.py" + ("" if base.endswith("fotos_nuevas") else f" '{base}'"))


if __name__ == "__main__":
    main()
