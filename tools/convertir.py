#!/usr/bin/env python3
"""
Conversor Kialo bebé: Excel -> productos.json
Uso:  python3 convertir.py            (lee Kialo_Catalogo.xlsx)
      python3 convertir.py otro.xlsx

Genera 'productos.json' en la raíz del sitio.
- FOTOS: varias por producto, separadas por coma (ej: boina.jpg, boina-2.jpg).
- COLORES: por NOMBRE en español (blanco, rojo, rosa...). Se validan; no hace falta saber hex.
"""
import json, sys, os, re, unicodedata
from openpyxl import load_workbook

CATEGORIAS = [
    {"id": "accesorios", "nombre": "Accesorios", "emoji": "🎀", "desc": "Boinas, lentes y más"},
    {"id": "ropa",       "nombre": "Ropa",       "emoji": "👕", "desc": "Suave y abrigadora"},
    {"id": "zapatillas", "nombre": "Zapatillas", "emoji": "👟", "desc": "Primeros pasitos"},
    {"id": "tematicas",  "nombre": "Temáticas",  "emoji": "🦊", "desc": "Para sus momentos especiales"},
]
IDS = {c["id"] for c in CATEGORIAS}

# Colores válidos (mismos nombres que entiende la web). Sirve solo para validar.
COLORES_VALIDOS = {
    "blanco","negro","gris","plomo","rojo","coral","rosa","rosado","fucsia","amarillo","mostaza",
    "naranja","naranjado","verde","verde agua","turquesa","celeste","azul","azul marino","morado",
    "lila","violeta","beige","crema","marron","marrón","cafe","café","dorado","plateado","vino"
}

def num(v):
    if v is None or v == "": return None
    try: return int(v) if float(v).is_integer() else float(v)
    except (ValueError, TypeError): return None

def lista(celda):
    if not celda: return []
    return [x.strip() for x in str(celda).split(",") if x.strip()]

def slug_codigo(nombre):
    """Código de pedido legible desde el nombre: primera palabra, sin acentos, MAYÚSCULAS.
    Ej: 'Boina con lazo' -> 'BOINA'. La unicidad se garantiza aparte con un sufijo numérico."""
    s = unicodedata.normalize("NFKD", str(nombre)).encode("ascii", "ignore").decode()
    s = re.sub(r"[^A-Za-z0-9 ]", "", s).upper().strip()
    palabras = s.split()
    return palabras[0] if palabras else "PROD"

def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "Kialo_Catalogo.xlsx"
    wb = load_workbook(xlsx, data_only=True)
    ws = wb["Productos"]

    # Mapea encabezados -> índice (tolera orden distinto)
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    def col(*nombres):
        for n in nombres:
            for i,h in enumerate(headers):
                if h.startswith(n): return i
        return None
    iNom=col("nombre"); iCat=col("categor"); iPre=col("precio"); iAnt=col("antes")
    iFot=col("foto"); iDesc=col("descrip"); iDet=col("detalle"); iTalla=col("talla")
    iMat=col("material"); iCol=col("color")

    productos, errores, avisos = [], [], []
    codigos_usados = set()
    for r in range(2, ws.max_row+1):
        fila = [c.value for c in ws[r]]
        def g(i): return fila[i] if (i is not None and i < len(fila)) else None
        nombre = (str(g(iNom)).strip() if g(iNom) else "")
        if not nombre: continue
        cat = (str(g(iCat)).strip().lower() if g(iCat) else "")
        if cat not in IDS:
            errores.append(f"Fila {r}: categoría '{cat}' inválida (usa: accesorios, ropa, zapatillas, tematicas)."); continue
        precio = num(g(iPre))
        if precio is None:
            errores.append(f"Fila {r} ('{nombre}'): falta el precio."); continue
        antes = num(g(iAnt))
        fotos = lista(g(iFot))
        if not fotos:
            avisos.append(f"Fila {r} ('{nombre}'): sin fotos.")
        colores = []
        for c in lista(g(iCol)):
            cl = c.lower()
            if cl in COLORES_VALIDOS: colores.append(cl)
            else: avisos.append(f"Fila {r} ('{nombre}'): color '{c}' no reconocido (se omitió).")
        # Código único y estable para el mensaje de WhatsApp (BOINA, BOINA2, ...)
        codigo = base = slug_codigo(nombre); n = 2
        while codigo in codigos_usados:
            codigo = f"{base}{n}"; n += 1
        codigos_usados.add(codigo)
        productos.append({
            "nombre": nombre, "codigo": codigo, "categoria": cat, "precio": precio,
            "antes": antes if (antes and antes > precio) else None,
            "fotos": fotos,
            "desc": (str(g(iDesc)).strip() if g(iDesc) else ""),
            "detalle": (str(g(iDet)).strip() if g(iDet) else ""),
            "talla": (str(g(iTalla)).strip() if g(iTalla) else ""),
            "material": (str(g(iMat)).strip() if g(iMat) else ""),
            "colores": colores,
        })

    tienda = {"nombre":"Kialo bebé","whatsapp":"51955105631","tiktok":"#","instagram":"#"}
    if "Tienda" in wb.sheetnames:
        for row in wb["Tienda"].iter_rows(min_row=2, values_only=True):
            campo=(str(row[0]).lower() if row[0] else ""); val=(str(row[1]).strip() if row[1] else "")
            if not val: continue
            if "nombre" in campo: tienda["nombre"]=val
            elif "whatsapp" in campo: tienda["whatsapp"]="".join(ch for ch in val if ch.isdigit())
            elif "tiktok" in campo: tienda["tiktok"]=val
            elif "instagram" in campo: tienda["instagram"]=val

    data={"tienda":tienda,"categorias":CATEGORIAS,"productos":productos}
    destino = os.path.join("..","productos.json") if os.path.basename(os.getcwd())=="tools" else "productos.json"
    with open(destino,"w",encoding="utf-8") as f:
        json.dump(data,f,ensure_ascii=False,indent=2)

    print(f"✅ productos.json generado con {len(productos)} producto(s)  ->  {os.path.abspath(destino)}")
    if avisos:
        print("\n🔔 Avisos:")
        for a in avisos: print("   - "+a)
    if errores:
        print("\n⚠️  Filas NO incluidas:")
        for e in errores: print("   - "+e)

if __name__ == "__main__":
    main()
